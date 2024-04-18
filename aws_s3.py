import io
import os
import boto3
import pandas as pd
from openpyxl import load_workbook
import google.generativeai as genai
from dotenv import load_dotenv
import time

load_dotenv()

API_KEY = os.getenv("API_KEY")
AWS_BDD_INPUT_BUCKET = os.getenv("aws_bdd_input_bucket")
AWS_BDD_OUTPUT_BUCKET = os.getenv("aws_bdd_output_bucket")
AWS_ARCHIVE_BUCKET = os.getenv("aws_bdd_archive_bucket")
AWS_ACCESS_KEY_ID = os.getenv("aws_access_key_id")
AWS_SECRET_ACCESS_KEY = os.getenv("aws_secret_access_key")
# AWS_SESSION_TOKEN = os.getenv("AWS_SESSION_TOKEN")
AWS_LOB_FILES = os.getenv("aws_lob_files")
AWS_TEST_OUTPUT_BUCKET = os.getenv("aws_test_output_bucket")

genai.configure(api_key=API_KEY)
generation_config = {
    "temperature": 0.9,
    "top_p": 1,
    "top_k": 1,
    "max_output_tokens": 2048,
}

safety_settings = [
    {
        "category": "HARM_CATEGORY_HARASSMENT",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_HATE_SPEECH",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_SEXUALLY_EXPLICIT",
        "threshold": "BLOCK_NONE"
    },
    {
        "category": "HARM_CATEGORY_DANGEROUS_CONTENT",
        "threshold": "BLOCK_NONE"
    },
]

model = genai.GenerativeModel(model_name="gemini-1.0-pro",
                              generation_config=generation_config, safety_settings=safety_settings)

s3_client = boto3.client(
    "s3",
    aws_access_key_id=AWS_ACCESS_KEY_ID,
    aws_secret_access_key=AWS_SECRET_ACCESS_KEY
    # aws_session_token=AWS_SESSION_TOKEN,
)


def upload_file_to_s3(username):
    try:
        file = f"./static/uploads/{username}_input.xlsx"
        s3_client.upload_file(file, AWS_BDD_INPUT_BUCKET, f'{username}_input.xlsx')
        return True
    except Exception as e:
        print(e)
        return False


def generate_bdd_from_jira(user_story):
    responses = []
    for story in user_story:
        convo = model.start_chat()
        convo.send_message("Generate BDD scenario in feature file format for the  user story " + story)
        response = convo.last.text
        responses.append([story, response])
    df1 = pd.DataFrame(responses)
    with io.StringIO() as csv_buffer:
        df1.to_csv(csv_buffer, index=False)
        ts = str(int(round(time.time())))
        response = s3_client.put_object(
            Bucket=AWS_BDD_OUTPUT_BUCKET, Key=f"output_{ts}.csv", Body=csv_buffer.getvalue()
        )
        status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")
        url = f"https://{AWS_BDD_OUTPUT_BUCKET}.s3.amazonaws.com/output_{ts}.csv"
        if status == 200:
            return url
        else:
            return None


def generate_bdd_scenario(username):
    s3_client_data = s3_client.get_object(Bucket=AWS_BDD_INPUT_BUCKET, Key=f'{username}_input.xlsx')
    contents = s3_client_data['Body'].read()  # your Excel's essence, pretty much a stream
    # Read in data_only mode to parse Excel after all formulae evaluated
    wb = load_workbook(filename=(io.BytesIO(contents)), data_only=True)
    sheet = wb.active
    responses = []
    for row in range(2, sheet.max_row + 1):
        # for row in range(1, sheet.max_row+1):
        prompt = sheet.cell(row, 1).value
        # Generate response
        convo = model.start_chat()
        convo.send_message("Generate BDD scenario in feature file format for the  user story " + prompt)
        response = convo.last.text
        # Save response
        responses.append(response)
        # print(prompt)
        # response = prompt
        # responses.append(response)
    df1 = pd.DataFrame(responses)
    with io.StringIO() as csv_buffer:
        df1.to_csv(csv_buffer, index=False)
        ts = str(int(round(time.time())))
        response = s3_client.put_object(
            Bucket=AWS_BDD_OUTPUT_BUCKET, Key=f"output_{ts}.csv", Body=csv_buffer.getvalue()
        )
        status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")
        s3 = boto3.resource('s3')
        s3.Object(AWS_ARCHIVE_BUCKET, f'{username}_input_{ts}.xlsx').copy_from(
            CopySource=f'{AWS_BDD_INPUT_BUCKET}/{username}_input.xlsx')
        s3.Object(AWS_BDD_INPUT_BUCKET, f'{username}_input.xlsx').delete()
        url = f"https://{AWS_BDD_OUTPUT_BUCKET}.s3.amazonaws.com/output_{ts}.csv"
        if status == 200:
            return url
        else:
            return None


def generate_test_data(lob, state, no_of_test_cases):
    s3_client_data = s3_client.get_object(Bucket=AWS_LOB_FILES, Key=f'{lob}.txt')
    contents = s3_client_data['Body'].read()  # Reading the txt file
    responses = []
    round_of_test_data = int(no_of_test_cases) // 10
    for test_cases_no in range(round_of_test_data + 1):
        # Generate response
        prompt = (f"Generate 10 test data for a {lob} policy according to the following criteria:\n"
                  f"include state {state} and {lob} for the line of business  using the following data\n"
                  + contents.decode('utf-8') + "\n in a csv format only.")
        # print(prompt)
        convo = model.start_chat()
        convo.send_message(prompt)
        response = convo.last.text
        # Save response
        if test_cases_no == 0:
            responses.append(response)
        else:
            response = "\n".join(response.split("\n")[1:])
            responses.append(response)
        # responses_bytes += response.encode('utf-8')
    responses_bytes = ("\n".join([response for response in responses])).encode('utf-8')
    ts = str(int(round(time.time())))
    response = s3_client.put_object(
        Bucket=AWS_TEST_OUTPUT_BUCKET, Key=f"{lob}_{ts}.csv", Body=responses_bytes
    )
    status = response.get("ResponseMetadata", {}).get("HTTPStatusCode")
    url = f"https://{AWS_TEST_OUTPUT_BUCKET}.s3.amazonaws.com/{lob}_{ts}.csv"
    if status == 200:
        return url
    else:
        return None
